import openpyxl
import tkinter as tk
from tkinter import filedialog

# Sélection du fichier par l'utilisateur
root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilename()

# Init workbook
wb_obj = openpyxl.load_workbook(file_path)

# Définir Sheet1
sheet1_obj = wb_obj.active

# Création Sheet2
sheet2_obj = wb_obj.create_sheet("Sheet2")

# Nombre de lignes
row = sheet1_obj.max_row

for i in range(row):
    if sheet1_obj.cell(row=i+13, column=5).value is not None: # +13 parce que le tableau commence à la 13ème ligne
        sheet1_obj.cell(row=i+13, column=5).value = float(sheet1_obj.cell(row=i+13, column=5).value) # Transforme les string des décimales en float pour la somme

for j in range(row):
    piece = sheet1_obj.cell(row=j+13, column=4).value
    aire = sheet1_obj.cell(row=j+13, column=5).value
    sheet2_obj.cell(row=1, column=j*2+1).value = piece # j * 2 + 1 pour les impairs
    sheet2_obj.cell(row=1, column=(j+1)*2).value = aire # j + 1 * 2 pour les pairs

sheet1_obj.cell(row=row+1, column=5).value = f"=SUM(E13:E{row})"

# Sauvegarder
wb_obj.save(file_path)